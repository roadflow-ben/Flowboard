import re
import streamlit as st
import pandas as pd
from datetime import date, datetime, timedelta, time
from io import BytesIO
from copy import copy as pycopy

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =========================================================
# Flowboard — MVP v0.3
# - Bible locked
# - Overdue = Dark Blue
# - Oldest cutoff first
# - Territory mapping (Option B)
# - Day focus territory to prevent ridiculous cross-region planning
# - Styled Excel export preserving existing colour coding
# =========================================================

st.set_page_config(page_title="Flowboard", page_icon="🧠", layout="wide")


# -----------------------------
# UI styling: calm green accents
# -----------------------------
st.markdown(
    """
    <style>
      /* Try to force Streamlit's primary colour to a calm green */
      :root { --primary-color: #2e7d32 !important; }
      /* Native checkbox tint (works if Streamlit renders real inputs) */
      input[type="checkbox"] { accent-color: #2e7d32 !important; }
      /* Streamlit checkbox/radio visuals often use SVG icons */
      div[data-testid="stCheckbox"] svg,
      div[data-testid="stRadio"] svg,
      div[data-testid="stToggle"] svg {
        color: #2e7d32 !important;
        fill: #2e7d32 !important;
      }
      /* BaseWeb (Streamlit) checkbox/toggle internals */
      label[data-baseweb="checkbox"] svg,
      label[data-baseweb="radio"] svg {
        color: #2e7d32 !important;
        fill: #2e7d32 !important;
      }
      /* When checked, the box background sometimes uses currentColor */
      label[data-baseweb="checkbox"] div[role="checkbox"],
      label[data-baseweb="radio"] div[role="radio"] {
        border-color: #2e7d32 !important;
      }
      /* Some builds use masks/strokes */
      div[data-testid="stCheckbox"] svg path,
      div[data-testid="stRadio"] svg path {
        stroke: #2e7d32 !important;
      }
    </style>
    """,
    unsafe_allow_html=True
)

WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
LOAD_MODES = ["Light", "Normal", "Heavy"]  # Heavy = +20%
LOAD_MULTIPLIER = {"Light": 0.85, "Normal": 1.00, "Heavy": 1.20}

EXPORT_SHEET_NAME = "Completed Schedule"
EXPORT_DATE_COL = "Survey_Date"
EXPORT_AMPM_COL = "am_pm"
EXPORT_ISO_WEEK_COL = "ISO_Week"


def monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())


def as_date(x):
    if pd.isna(x):
        return None
    if isinstance(x, date) and not isinstance(x, datetime):
        return x
    if isinstance(x, datetime):
        return x.date()
    try:
        return pd.to_datetime(x).date()
    except Exception:
        return None


def pick_col(cols, candidates):
    """Return first matching column (case-insensitive contains). Works even if Excel has date headers."""
    cols_str = [str(c) for c in cols]
    cols_lower = {c.lower(): orig for c, orig in zip(cols_str, cols)}
    for cand in candidates:
        for c_str, orig in zip(cols_str, cols):
            if cand in c_str.lower():
                return orig
    for cand in candidates:
        if cand in cols_lower:
            return cols_lower[cand]
    return None


def normalize_address(row, number_col, street_col, suburb_col, city_col):
    parts = []
    if number_col and pd.notna(row.get(number_col, None)):
        parts.append(str(row[number_col]).strip())
    if street_col and pd.notna(row.get(street_col, None)):
        parts.append(str(row[street_col]).strip())
    addr = " ".join(parts).strip()

    loc_parts = []
    if suburb_col and pd.notna(row.get(suburb_col, None)):
        loc_parts.append(str(row[suburb_col]).strip())
    if city_col and pd.notna(row.get(city_col, None)):
        loc_parts.append(str(row[city_col]).strip())
    loc = ", ".join([p for p in loc_parts if p])

    return addr if not loc else f"{addr} — {loc}"


def cutoff_date(target_date: date):
    return (target_date + timedelta(days=30)) if target_date else None


def urgency_band(target_date: date, week_start: date):
    """
    Bible + fix:
    - Dark Blue if last-chance week OR already overdue (cutoff passed).
    - Light Blue if 1-2 weeks before last-chance week.
    """
    if not target_date:
        return "Flexible"

    window_close = cutoff_date(target_date)
    last_week_start = monday_of_week(window_close)

    # overdue or last-chance week
    if week_start >= last_week_start:
        return "Dark Blue"

    if week_start in (last_week_start - timedelta(days=7), last_week_start - timedelta(days=14)):
        return "Light Blue"

    return "Flexible"


def futile_rank(status_val):
    if status_val is None or (isinstance(status_val, float) and pd.isna(status_val)):
        return 0
    s = str(status_val).strip().lower()
    if "futile 2" in s or "futile2" in s:
        return 2
    if "futile 1" in s or "futile1" in s:
        return 1
    return 0


def estimate_minutes(bedrooms, inspection_type=None):
    try:
        b = int(float(bedrooms))
    except Exception:
        b = None

    if b is None:
        base = 15
    elif b <= 1:
        base = 7
    elif b <= 3:
        base = 15
    else:
        base = 40

    if inspection_type:
        t = str(inspection_type).lower()
        if "plus" in t or "full" in t or "condition" in t:
            base = int(base * 1.35)

    return base


def session_capacity_minutes(time_mode, global_times, day_override_times, day_name, session_name, load_mode):
    times = day_override_times.get(day_name) or global_times

    if time_mode == "Inspection window":
        start_t = times["start_first"]
        end_t = times["latest_arrival_last"]
    else:
        start_t = times["depart_depot"]
        end_t = times["return_depot"]

    if not (start_t and end_t):
        base_minutes = 240
    else:
        dt0 = datetime.combine(date.today(), start_t)
        dt1 = datetime.combine(date.today(), end_t)
        base_minutes = max(0, int((dt1 - dt0).total_seconds() // 60))

    sess = int(base_minutes * 0.55) if session_name == "AM" else int(base_minutes * 0.45)
    sess = int(sess * LOAD_MULTIPLIER[load_mode])
    sess = int(sess * 0.90)  # safety buffer
    return max(60, sess)


# -----------------------------
# Excel styled export
# -----------------------------
def find_or_add_column(ws, header_name: str) -> int:
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        val = ws.cell(row=1, column=c).value
        if str(val).strip() == header_name:
            return c

    new_col = max_col + 1
    hdr_cell = ws.cell(row=1, column=new_col)
    hdr_cell.value = header_name

    if max_col >= 1:
        prev = ws.cell(row=1, column=max_col)
        hdr_cell._style = pycopy(prev._style)
        hdr_cell.font = pycopy(prev.font)
        hdr_cell.fill = pycopy(prev.fill)
        hdr_cell.border = pycopy(prev.border)
        hdr_cell.alignment = pycopy(prev.alignment)
        hdr_cell.number_format = prev.number_format
        hdr_cell.protection = pycopy(prev.protection)

    ws.column_dimensions[get_column_letter(new_col)].width = max(
        14, ws.column_dimensions[get_column_letter(max_col)].width or 14
    )
    return new_col


def copy_row_with_styles(ws_src, ws_dst, src_row: int, dst_row: int, max_col: int):
    for c in range(1, max_col + 1):
        cell_src = ws_src.cell(row=src_row, column=c)
        cell_dst = ws_dst.cell(row=dst_row, column=c)

        cell_dst.value = cell_src.value

        cell_dst._style = pycopy(cell_src._style)
        cell_dst.font = pycopy(cell_src.font)
        cell_dst.fill = pycopy(cell_src.fill)
        cell_dst.border = pycopy(cell_src.border)
        cell_dst.alignment = pycopy(cell_src.alignment)
        cell_dst.number_format = cell_src.number_format
        cell_dst.protection = pycopy(cell_src.protection)

        if cell_src.comment:
            cell_dst.comment = pycopy(cell_src.comment)


def build_styled_completed_workbook(original_bytes: bytes, plan_df: pd.DataFrame) -> bytes:
    wb = load_workbook(BytesIO(original_bytes))
    ws_src = wb.active

    date_col_idx = find_or_add_column(ws_src, EXPORT_DATE_COL)
    ampm_col_idx = find_or_add_column(ws_src, EXPORT_AMPM_COL)
    iso_week_col_idx = find_or_add_column(ws_src, EXPORT_ISO_WEEK_COL)

    if plan_df is None or plan_df.empty:
        if EXPORT_SHEET_NAME in wb.sheetnames:
            del wb[EXPORT_SHEET_NAME]
        ws_out = wb.create_sheet(EXPORT_SHEET_NAME)

        max_col = ws_src.max_column
        max_row = ws_src.max_row
        for r in range(1, max_row + 1):
            copy_row_with_styles(ws_src, ws_out, r, r, max_col)

        out = BytesIO()
        wb.active = wb.sheetnames.index(EXPORT_SHEET_NAME)
        wb.save(out)
        return out.getvalue()

    # write planned fields into source sheet
    for _, r in plan_df.iterrows():
        excel_row = int(r["_excel_row"])
        pdate = r.get("_planned_date")
        psess = r.get("_planned_session")
        if pd.notna(pdate) and pdate is not None:
            ws_src.cell(row=excel_row, column=date_col_idx).value = pdate
            try:
                ws_src.cell(row=excel_row, column=iso_week_col_idx).value = int(as_date(pdate).isocalendar()[1])
            except Exception:
                ws_src.cell(row=excel_row, column=iso_week_col_idx).value = None
        if psess:
            ws_src.cell(row=excel_row, column=ampm_col_idx).value = str(psess)

    plan_sorted = plan_df.copy()
    plan_sorted["_planned_date_sort"] = plan_sorted["_planned_date"].apply(lambda x: x if isinstance(x, date) else as_date(x))
    plan_sorted["_sess_sort"] = plan_sorted["_planned_session"].map({"AM": 0, "PM": 1}).fillna(9)
    plan_sorted = plan_sorted.sort_values(by=["_planned_date_sort", "_sess_sort", "_planned_seq"], ascending=True)

    scheduled_rows = [int(x) for x in plan_sorted["_excel_row"].tolist()]
    scheduled_set = set(scheduled_rows)

    if EXPORT_SHEET_NAME in wb.sheetnames:
        del wb[EXPORT_SHEET_NAME]
    ws_out = wb.create_sheet(EXPORT_SHEET_NAME)

    max_col = ws_src.max_column
    max_row = ws_src.max_row

    copy_row_with_styles(ws_src, ws_out, 1, 1, max_col)
    out_row = 2

    for src_r in scheduled_rows:
        if 2 <= src_r <= max_row:
            copy_row_with_styles(ws_src, ws_out, src_r, out_row, max_col)
            out_row += 1

    for src_r in range(2, max_row + 1):
        if src_r not in scheduled_set:
            copy_row_with_styles(ws_src, ws_out, src_r, out_row, max_col)
            out_row += 1

    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        ws_out.column_dimensions[col_letter].width = ws_src.column_dimensions[col_letter].width

        # Ensure the export OPENS on the sorted sheet
    try:
        ws_out = wb[EXPORT_SHEET_NAME]
        wb.active = wb.sheetnames.index(EXPORT_SHEET_NAME)
        # Move Completed Schedule to be the first tab (optional but helps clarity)
        wb._sheets.remove(ws_out)
        wb._sheets.insert(0, ws_out)
        wb.active = 0
    except Exception:
        pass

    out = BytesIO()
    wb.save(out)
    return out.getvalue()



# -----------------------------
# State init
# -----------------------------
if "view" not in st.session_state:
    st.session_state.view = "setup"
if "df" not in st.session_state:
    st.session_state.df = None
if "original_bytes" not in st.session_state:
    st.session_state.original_bytes = None
if "colmap" not in st.session_state:
    st.session_state.colmap = {}
if "plan" not in st.session_state:
    st.session_state.plan = None
if "plan_df" not in st.session_state:
    st.session_state.plan_df = None


# -----------------------------
# Header
# -----------------------------
st.markdown(
    """
    <div style="margin-top:6px;">
      <div style="display:flex;align-items:baseline;gap:14px;flex-wrap:wrap;">
        <div style="font-size:44px;font-weight:850;line-height:1;">Flowboard</div>
        <div style="font-size:14px;opacity:0.75;font-weight:600;">Clarity before commitment.</div>
      </div>

      <div style="margin-top:14px;padding:14px 16px;border:1px solid rgba(0,0,0,0.06);border-radius:14px;background:rgba(46,125,50,0.06);max-width:900px;">
        <div style="font-size:18px;font-weight:800;line-height:1.25;">
          Plan a realistic week of site visits — one week at a time — before backlog becomes pressure.
        </div>

        <div style="margin-top:10px;font-size:14px;line-height:1.55;opacity:0.9;">
          Flowboard turns a daunting list of future site visits into a clear, structured plan —
          removing the chaos, respecting your available time, chosen working days,
          and using light location grouping to keep each day achievable.
        </div>

        <div style="margin-top:10px;font-size:14px;font-style:italic;opacity:0.85;">
          Overflow is expected. That’s the point.
        </div>
      </div>
    </div>
    """,
    unsafe_allow_html=True
)

# -----------------------------
# Sidebar: Week Setup
# -----------------------------
with st.sidebar:
    st.subheader("Week Setup")

    uploaded = st.file_uploader("Import Backlog (Excel)", type=["xlsx", "xls"])
    if uploaded is not None:
        st.session_state.original_bytes = uploaded.getvalue()
        st.session_state.df = pd.read_excel(BytesIO(st.session_state.original_bytes))

    df = st.session_state.df
    if df is None:
        st.info("Upload an Excel backlog to begin.")
        st.stop()

    cols = list(df.columns)

    # Week selection (ISO week)
    st.markdown("**Week starting**")
    _picked = st.date_input(
        "Week starting",
        value=monday_of_week(date.today()),
        label_visibility="collapsed",
        key="week_start_date",
    )
    week_start = monday_of_week(_picked)
    iso_year, iso_week, _ = week_start.isocalendar()
    st.caption(f"ISO Wk {iso_week} ({iso_year}) — Mon {week_start.strftime('%d/%m/%Y')}")

    # Working days (always offer all 7 days)
    st.caption("Working days:")
    active_days = {}
    default_on = {d: (d in WEEKDAYS[:5]) for d in WEEKDAYS}  # default Mon-Fri on
    cols_days = st.columns(7)
    for i, d in enumerate(WEEKDAYS):
        with cols_days[i]:
            active_days[d] = st.checkbox("", value=default_on[d], key=f"day_{d}")
            st.markdown(
                f"<div style='text-align:center;font-size:12px;opacity:0.75'>{d[:3]}</div>",
                unsafe_allow_html=True
            )


    # Time bounds
    st.divider()
    st.caption("Time bounds (global, with per-day overrides):")
    time_mode = st.radio(
        "Mode",
        ["Inspection window", "Depot window"],
        horizontal=True,
        label_visibility="collapsed",
        key="time_mode",
    )

    if time_mode == "Inspection window":
        start_first = st.time_input("Start 1st inspection", value=time(8, 30), key="start_first")
        latest_arrival_last = st.time_input(
            "Latest arrival @ last inspection",
            value=time(15, 30),
            key="latest_arrival_last",
        )
        global_times = {
            "start_first": start_first,
            "latest_arrival_last": latest_arrival_last,
            "depart_depot": None,
            "return_depot": None,
        }
    else:
        depart_depot = st.time_input("Depart depot", value=time(8, 0), key="depart_depot")
        return_depot = st.time_input("Return depot", value=time(16, 30), key="return_depot")
        global_times = {
            "start_first": None,
            "latest_arrival_last": None,
            "depart_depot": depart_depot,
            "return_depot": return_depot,
        }

    # Sessions + load + overrides + focus per day
    st.divider()
    st.caption("Session loading & day parameters")

    day_override_times = {}
    day_sessions = {}

    for d in WEEKDAYS:
        if not active_days.get(d, False):
            continue

        with st.expander(d, expanded=False):
            am_enabled = st.checkbox("AM enabled", value=True, key=f"{d}_am_on")
            pm_enabled = st.checkbox("PM enabled", value=True, key=f"{d}_pm_on")

            am_load = st.selectbox("AM load", LOAD_MODES, index=1, key=f"{d}_am_load")
            pm_load = st.selectbox("PM load", LOAD_MODES, index=1, key=f"{d}_pm_load")

            override = st.checkbox("Override times for this day", value=False, key=f"{d}_override")
            if override:
                if time_mode == "Inspection window":
                    o_start = st.time_input(
                        "Start 1st inspection (override)",
                        value=global_times["start_first"],
                        key=f"{d}_o_start",
                    )
                    o_last = st.time_input(
                        "Latest arrival last (override)",
                        value=global_times["latest_arrival_last"],
                        key=f"{d}_o_last",
                    )
                    day_override_times[d] = {
                        "start_first": o_start,
                        "latest_arrival_last": o_last,
                        "depart_depot": None,
                        "return_depot": None,
                    }
                else:
                    o_dep = st.time_input(
                        "Depart depot (override)",
                        value=global_times["depart_depot"],
                        key=f"{d}_o_dep",
                    )
                    o_ret = st.time_input(
                        "Return depot (override)",
                        value=global_times["return_depot"],
                        key=f"{d}_o_ret",
                    )
                    day_override_times[d] = {
                        "start_first": None,
                        "latest_arrival_last": None,
                        "depart_depot": o_dep,
                        "return_depot": o_ret,
                    }

            day_sessions[d] = {
                "AM": {"enabled": am_enabled, "load": am_load},
                "PM": {"enabled": pm_enabled, "load": pm_load},
                "focus": None,
            }

    st.divider()
    go = st.button("Go ahead, PLAN my day", type="primary", use_container_width=True)


# -----------------------------
# Main panel: mapping + overview
# -----------------------------
df = st.session_state.df
cols = list(df.columns)

# Auto-detect mapping columns
auto_target = pick_col(cols, ["target_date", "target date", "due", "target"])
auto_status = pick_col(cols, ["status", "survey_status", "survey status", "state"])
auto_bed = pick_col(cols, ["bdrm", "bed", "bedroom", "bdrm_no", "bdrm no"])
auto_type = pick_col(cols, ["inspection type", "type", "visit type"])
auto_ref = pick_col(cols, ["reference", "property_reference", "property reference", "id"])
auto_street = pick_col(cols, ["street"])
auto_number = pick_col(cols, ["number", "street number", "no."])
auto_suburb = pick_col(cols, ["suburb"])
auto_city = pick_col(cols, ["city", "town", "region", "area"])

with st.expander("Data mapping (optional)", expanded=False):
    st.caption("Flowboard is input-format agnostic. These defaults are detected; change if needed.")
    col_target = st.selectbox("Target date column", ["(none)"] + cols, index=(["(none)"] + cols).index(auto_target) if auto_target in cols else 0)
    col_status = st.selectbox("Status column (e.g. Futile 1/2)", ["(none)"] + cols, index=(["(none)"] + cols).index(auto_status) if auto_status in cols else 0)
    col_bed = st.selectbox("Bedrooms column", ["(none)"] + cols, index=(["(none)"] + cols).index(auto_bed) if auto_bed in cols else 0)
    col_type = st.selectbox("Inspection type column", ["(none)"] + cols, index=(["(none)"] + cols).index(auto_type) if auto_type in cols else 0)
    col_ref = st.selectbox("Reference/ID column", ["(none)"] + cols, index=(["(none)"] + cols).index(auto_ref) if auto_ref in cols else 0)
    col_number = st.selectbox("Street number column", ["(none)"] + cols, index=(["(none)"] + cols).index(auto_number) if auto_number in cols else 0)
    col_street = st.selectbox("Street name column", ["(none)"] + cols, index=(["(none)"] + cols).index(auto_street) if auto_street in cols else 0)
    col_suburb = st.selectbox("Suburb column", ["(none)"] + cols, index=(["(none)"] + cols).index(auto_suburb) if auto_suburb in cols else 0)
    col_city = st.selectbox("City/Area column", ["(none)"] + cols, index=(["(none)"] + cols).index(auto_city) if auto_city in cols else 0)

st.session_state.colmap = {
    "target": None if col_target == "(none)" else col_target,
    "status": None if col_status == "(none)" else col_status,
    "bed": None if col_bed == "(none)" else col_bed,
    "type": None if col_type == "(none)" else col_type,
    "ref": None if col_ref == "(none)" else col_ref,
    "number": None if col_number == "(none)" else col_number,
    "street": None if col_street == "(none)" else col_street,
    "suburb": None if col_suburb == "(none)" else col_suburb,
    "city": None if col_city == "(none)" else col_city,
}
cm = st.session_state.colmap

st.subheader("Planning Overview")

# -----------------------------
# Area grouping (MVP)
# -----------------------------
# We use a single "area" grouping column for planning. By default this is Suburb (best),
# otherwise City/Town/Region/Area if available.
col_geo = cm.get("suburb") if cm.get("suburb") in df.columns else None
if col_geo is None:
    col_geo = cm.get("city") if cm.get("city") in df.columns else None

# Build derived dataframe
df_work = df.copy()
df_work["_excel_row"] = df_work.reset_index().index + 2

# target + urgency + cutoff
if cm["target"]:
    df_work["_target_date"] = df_work[cm["target"]].apply(as_date)
else:
    df_work["_target_date"] = None

df_work["_cutoff_date"] = df_work["_target_date"].apply(cutoff_date)
df_work["_urgency"] = df_work["_target_date"].apply(lambda td: urgency_band(td, week_start))

# label
df_work["_label"] = df_work.apply(
    lambda r: normalize_address(r, cm["number"], cm["street"], cm["suburb"], cm["city"]),
    axis=1
)

# estimates + futile rank
df_work["_mins"] = df_work.apply(
    lambda r: estimate_minutes(r.get(cm["bed"]) if cm["bed"] else None,
                               r.get(cm["type"]) if cm["type"] else None),
    axis=1
)
df_work["_futile_rank"] = df_work[cm["status"]].apply(futile_rank) if cm["status"] else 0

# territory from mapping
if col_geo and col_geo in df_work.columns:
    geo_series = df_work[col_geo].fillna("").astype(str).str.strip()
    df_work["_territory"] = geo_series.apply(lambda v: v if v else "Unknown")
else:
    df_work["_territory"] = "Unknown"


# -----------------------------
# Daily focus + per-day area availability (collapsible matrix)
# -----------------------------
st.subheader("Daily focus")
st.caption("Auto chooses the best Area for each day. Use Focus to override. Use the availability matrix to exclude Areas on specific days.")

areas = sorted(df_work["_territory"].fillna("Unknown").astype(str).unique().tolist())
active_day_list = [d for d in WEEKDAYS if st.session_state.get(f"day_{d}", False)]

# Day labels row stays visible even when matrix is collapsed (only active days)
if active_day_list:
    label_cols = st.columns(1 + len(active_day_list))
    with label_cols[0]:
        st.markdown("**Area**")
    for i, dlab in enumerate(active_day_list, start=1):
        with label_cols[i]:
            st.markdown(f"**{dlab[:3]}**")
else:
    st.info("No working days selected in the sidebar.")

# Ensure matrix state exists (default: all Areas available every day)
if "area_day_allowed" not in st.session_state:
    st.session_state.area_day_allowed = {d: {a: True for a in areas} for d in WEEKDAYS}
else:
    mat = st.session_state.area_day_allowed
    for d in WEEKDAYS:
        if d not in mat:
            mat[d] = {}
        # drop removed areas
        for old_a in list(mat[d].keys()):
            if old_a not in areas:
                del mat[d][old_a]
        # add new areas
        for a in areas:
            if a not in mat[d]:
                mat[d][a] = True
    st.session_state.area_day_allowed = mat

with st.expander("Area availability matrix (expand to include/exclude Areas per day)", expanded=False):
    st.caption("Untick an Area on a specific day to prevent Auto (and scheduling) from using it that day. Re-tick to re-enable.")
    if not areas:
        st.caption("No Areas detected.")
    else:
        for a in areas:
            row_cols = st.columns([2] + [1]*len(active_day_list))
            with row_cols[0]:
                st.write(a)
            for j, d in enumerate(active_day_list):
                key = f"allow::{a}::{d}"
                if key not in st.session_state:
                    st.session_state[key] = bool(st.session_state.area_day_allowed.get(d, {}).get(a, True))
                with row_cols[j + 1]:
                    st.checkbox("", key=key)
                st.session_state.area_day_allowed[d][a] = bool(st.session_state[key])

# Build day->allowed Areas map for the planner (only for active days)
day_allowed = {}
for d in active_day_list:
    allowed = {a for a in areas if st.session_state.area_day_allowed.get(d, {}).get(a, True)}
    day_allowed[d] = allowed

# Day Focus Area: if set, that day will schedule only jobs from that Area.
day_focus = {}
if active_day_list:
    cols_focus = st.columns(len(active_day_list))
    for i, d in enumerate(active_day_list):
        with cols_focus[i]:
            options = ["(auto)"] + areas
            k = f"focus_{d}"
            prev = st.session_state.get(k, "(auto)")
            if prev not in options:
                st.session_state[k] = "(auto)"
            day_focus[d] = st.selectbox(f"{d[:3]}", options, index=0, key=k)
else:
    day_focus = {}

# Overview metrics
c1, c2, c3, c4 = st.columns(4)
c1.metric("Dark Blue (Must this week)", int((df_work["_urgency"] == "Dark Blue").sum()))
c2.metric("Light Blue (Warning band)", int((df_work["_urgency"] == "Light Blue").sum()))
c3.metric("Areas detected", len(areas))
c4.metric("Flexible backlog", int((df_work["_urgency"] == "Flexible").sum()))

# Quick territory workload view (top 10)
t_counts = df_work["_territory"].value_counts().head(10).reset_index()
t_counts.columns = ["Area", "Jobs"]
st.write("**Workload by area (top 10)**")
st.dataframe(t_counts, use_container_width=True, hide_index=True)


# -----------------------------
# Planning Engine: territory-aware + day focus
# -----------------------------
def build_week_plan(df_in: pd.DataFrame, week_start: date, active_days, day_sessions, time_mode, global_times, day_override_times, day_focus, day_allowed=None):
    jobs = df_in.copy()

    # -----------------------------
    # Priority hierarchy (lower = more urgent)
    # -----------------------------
    urgency_order = {"Dark Blue": 0, "Light Blue": 1, "Flexible": 2}
    jobs["_urg_order"] = jobs["_urgency"].map(urgency_order).fillna(2).astype(int)

    # Tie-breakers / sorts
    jobs["_dark_tie"] = jobs.apply(lambda r: r.get("_futile_rank", 0) if r.get("_urgency") == "Dark Blue" else 0, axis=1)
    jobs["_cutoff_sort"] = jobs["_cutoff_date"].fillna(date.max)

    # Light geo grouping key (still helpful inside territory)
    geo_key_cols = []
    if cm.get("street"):
        geo_key_cols.append(cm["street"])
    if geo_key_cols and all(c in jobs.columns for c in geo_key_cols):
        jobs["_geo_key"] = jobs[geo_key_cols].astype(str).agg(" | ".join, axis=1)
    else:
        jobs["_geo_key"] = "Unknown"

    # Sort primarily by urgency + cutoff + futile, then territory, then street
    jobs = jobs.sort_values(
        by=["_urg_order", "_cutoff_sort", "_dark_tie", "_territory", "_geo_key", "_mins"],
        ascending=[True, True, True, True, True, True]
    ).reset_index(drop=True)

    # -----------------------------
    # Buckets init
    # -----------------------------
    buckets = {}
    for d in WEEKDAYS:
        if not active_days.get(d, False):
            continue
        buckets[d] = {"AM": [], "PM": []}

    remaining = jobs.to_dict(orient="records")

    # -----------------------------
    # Cluster key helper (conservative)
    # -----------------------------
    def derive_cluster_key(job: dict) -> str:
        label = str(job.get("_label", "") or "").strip().lower()
        if not label:
            return f"geo|{str(job.get('_geo_key','Unknown')).strip().lower()}"

        label = re.sub(r"^(unit|apt|apartment|flat)\s*\w+\s*,\s*", "", label)
        label = re.sub(r"^[a-z0-9]+\s*/\s*", "", label)

        m = re.match(r"^(\d+[a-z]?)\s+([a-z\s]+?)\s+(ave|avenue|rd|road|st|street|cres|crescent|pl|place|dr|drive|tce|terrace|ln|lane)\b", label)
        if m:
            num = m.group(1)
            street = re.sub(r"\s+", " ", m.group(2).strip())
            st_type = m.group(3)
            return f"bldg|{num}|{street}|{st_type}"

        return f"geo|{str(job.get('_geo_key','Unknown')).strip().lower()}"

    # -----------------------------
    # Territory choice (existing behaviour)
    # -----------------------------
    def choose_auto_territory(rem_list, allowed_terr=None):
        """Pick the territory with the most urgent weight remaining (Dark > Light > Flexible)."""
        if not rem_list:
            return None
        score = {}
        for r in rem_list:
            terr = str(r.get("_territory", "Unknown"))
            if allowed_terr is not None and terr not in allowed_terr:
                continue
            urg = r.get("_urgency", "Flexible")
            w = 100 if urg == "Dark Blue" else 10 if urg == "Light Blue" else 1
            score[terr] = score.get(terr, 0) + w
        return max(score.items(), key=lambda kv: kv[1])[0] if score else None

    def pop_first_matching(predicate):
        """Pop first item in remaining that matches predicate."""
        for i, item in enumerate(remaining):
            if predicate(item):
                return remaining.pop(i)
        return None

    def peek_any(predicate):
        """Check if any remaining item matches predicate."""
        return any(predicate(x) for x in remaining)

    # -----------------------------
    # Core scheduling loop
    # -----------------------------
    for d in [wd for wd in WEEKDAYS if active_days.get(wd, False)]:
        allowed_today = None
        if day_allowed is not None:
            allowed_today = set(day_allowed.get(d, []))
            if not allowed_today:
                allowed_today = None

        focus = day_focus.get(d, "(auto)")
        if focus is not None and focus != "(auto)" and allowed_today is not None and str(focus) not in allowed_today:
            focus = "(auto)"
        focus_terr = None if (focus is None or focus == "(auto)") else str(focus)

        if focus_terr is None:
            focus_terr = choose_auto_territory(remaining, allowed_today)

        if focus_terr is None:
            continue

        for sess in ["AM", "PM"]:
            if not day_sessions[d][sess]["enabled"]:
                continue

            load = day_sessions[d][sess]["load"]
            budget = session_capacity_minutes(time_mode, global_times, day_override_times, d, sess, load)

            used = 0
            picked = []

            while True:
                tier_in_terr = None
                for tier_name in ["Dark Blue", "Light Blue", "Flexible"]:
                    if peek_any(lambda x, t=tier_name: str(x.get("_territory", "Unknown")) == focus_terr and x.get("_urgency") == t):
                        tier_in_terr = tier_name
                        break

                if tier_in_terr is None:
                    break

                anchor = pop_first_matching(lambda x: str(x.get("_territory", "Unknown")) == focus_terr and x.get("_urgency") == tier_in_terr)
                if not anchor:
                    break

                anchor_m = int(anchor.get("_mins", 15))
                if used + anchor_m > int(budget * 1.10):
                    remaining.insert(0, anchor)
                    break

                ck = derive_cluster_key(anchor)
                batch = [anchor]
                batch_minutes = anchor_m

                def pop_same_tier_same_cluster():
                    return pop_first_matching(
                        lambda x: str(x.get("_territory", "Unknown")) == focus_terr
                        and x.get("_urgency") == tier_in_terr
                        and derive_cluster_key(x) == ck
                    )

                while len(batch) < 3:
                    nxt = pop_same_tier_same_cluster()
                    if not nxt:
                        break
                    m = int(nxt.get("_mins", 15))
                    if used + batch_minutes + m <= int(budget * 1.10):
                        batch.append(nxt)
                        batch_minutes += m
                    else:
                        remaining.insert(0, nxt)
                        break

                if len(batch) < 3:
                    if tier_in_terr == "Dark Blue":
                        lower_tiers = ["Flexible", "Light Blue"]
                    elif tier_in_terr == "Light Blue":
                        lower_tiers = ["Flexible"]
                    else:
                        lower_tiers = []

                    def would_starve_same_tier_if_add(extra_minutes: int) -> bool:
                        remaining_budget_after = int(budget * 1.10) - (used + batch_minutes + extra_minutes)
                        if remaining_budget_after <= 0:
                            return True
                        mins_same_tier = [
                            int(x.get("_mins", 15))
                            for x in remaining
                            if str(x.get("_territory", "Unknown")) == focus_terr and x.get("_urgency") == tier_in_terr
                        ]
                        if not mins_same_tier:
                            return False
                        return remaining_budget_after < min(mins_same_tier)

                    for lt in lower_tiers:
                        while len(batch) < 3:
                            pad = pop_first_matching(
                                lambda x, lt=lt: str(x.get("_territory", "Unknown")) == focus_terr
                                and x.get("_urgency") == lt
                                and derive_cluster_key(x) == ck
                            )
                            if not pad:
                                break

                            m = int(pad.get("_mins", 15))
                            if used + batch_minutes + m > int(budget * 1.10):
                                remaining.insert(0, pad)
                                break

                            if would_starve_same_tier_if_add(m):
                                remaining.insert(0, pad)
                                break

                            batch.append(pad)
                            batch_minutes += m

                        if len(batch) >= 3:
                            break

                for item in batch:
                    picked.append(item)
                used += batch_minutes

            for i, job in enumerate(picked, start=1):
                job["_planned_day"] = d
                job["_planned_date"] = week_start + timedelta(days=WEEKDAYS.index(d))
                job["_planned_session"] = sess
                job["_planned_seq"] = i

            buckets[d][sess] = picked

    planned_rows = []
    for d, sessions in buckets.items():
        for sess, items in sessions.items():
            planned_rows.extend(items)

    plan_df = pd.DataFrame(planned_rows) if planned_rows else pd.DataFrame()
    return buckets, plan_df, remaining


# -----------------------------
# GO button
# -----------------------------
if go:
    try:
        st.toast("Planning initiated. Hasta la vista, crazy. Removing the crazy from your work week.")
    except Exception:
        st.success("Planning initiated. Hasta la vista, crazy. Removing the crazy from your work week.")

    act = {d: st.session_state.get(f"day_{d}", False) for d in WEEKDAYS}
    sessions = {}
    for d in WEEKDAYS:
        if not act.get(d, False):
            continue
        sessions[d] = {
            "AM": {"enabled": st.session_state.get(f"{d}_am_on", True), "load": st.session_state.get(f"{d}_am_load", "Normal")},
            "PM": {"enabled": st.session_state.get(f"{d}_pm_on", True), "load": st.session_state.get(f"{d}_pm_load", "Normal")},
        }

    for d in list(day_focus.keys()):
        allowed_today = set(day_allowed.get(d, [])) if isinstance(day_allowed, dict) else set()
        if allowed_today and day_focus.get(d) not in ("(auto)", None) and day_focus[d] not in allowed_today:
            day_focus[d] = "(auto)"

    buckets, plan_df, remaining = build_week_plan(
        df_work, week_start, act, sessions, time_mode, global_times, day_override_times, day_focus, day_allowed
    )

    st.session_state.plan = {
        "week_start": week_start,
        "active_days": act,
        "day_sessions": sessions,
        "time_mode": time_mode,
        "global_times": global_times,
        "day_override_times": day_override_times,
        "buckets": buckets,
        "remaining": remaining,
        "day_focus": day_focus,
        "day_allowed": day_allowed,
    }
    st.session_state.plan_df = plan_df
    st.session_state.view = "review"


# -----------------------------
# Review Screen
# -----------------------------
URGENCY_COLORS = {"Dark Blue": "#1f4cff", "Light Blue": "#5aa9ff", "Flexible": "#9aa3af"}


def render_job(job):
    urg = job.get("_urgency", "Flexible")
    col = URGENCY_COLORS.get(urg, "#9aa3af")
    seq = job.get("_planned_seq", "")
    label = job.get("_label", "Unknown address")
    mins = job.get("_mins", 0)
    terr = job.get("_territory", "Unknown")
    return f"""
    <div style="display:flex;align-items:center;gap:10px;padding:6px 8px;border-bottom:1px dashed #e5e7eb;">
      <div style="width:26px;height:26px;border-radius:6px;background:{col};color:white;display:flex;align-items:center;justify-content:center;font-weight:800;">{seq}</div>
      <div style="flex:1;">
        <div style="font-weight:700;">{label}</div>
        <div style="font-size:12px;opacity:0.7;">{urg} • {terr} • est {mins} mins</div>
      </div>
    </div>
    """


if st.session_state.view == "review" and st.session_state.plan is not None:
    st.divider()
    plan = st.session_state.plan
    week_start = plan["week_start"]
    plan_df = st.session_state.plan_df

    h1, h2, h3 = st.columns([2, 1, 1])

    with h1:
        st.subheader("Weekly Plan Review")
        st.caption(
            f"Week starting: {week_start.strftime('%a %d %b %Y')} • Plans are conservative. Overflow is expected."
        )

    with h2:
        if st.button("Back to Week Setup", use_container_width=True):
            st.session_state.view = "setup"

    with h3:
        if st.session_state.original_bytes is not None:
            # --- ensure export order: Survey_Date, then am_pm, then stop sequence ---
            plan_df_export = plan_df.copy()

            if not plan_df_export.empty:
                date_col = "Survey_Date" if "Survey_Date" in plan_df_export.columns else "_planned_date"
                if date_col in plan_df_export.columns:
                    plan_df_export[date_col] = pd.to_datetime(plan_df_export[date_col], errors="coerce")

                sess_col = "am_pm" if "am_pm" in plan_df_export.columns else "_planned_session"
                sess_order = {"AM": 0, "PM": 1}
                plan_df_export["_sess_sort"] = (
                    plan_df_export.get(sess_col, "")
                    .astype(str)
                    .map(sess_order)
                    .fillna(9)
                    .astype(int)
                )

                sort_cols = [c for c in [date_col, "_sess_sort", "_planned_seq"] if c in plan_df_export.columns]
                plan_df_export = (
                    plan_df_export.sort_values(
                        by=sort_cols,
                        ascending=[True] * len(sort_cols),
                        kind="mergesort",
                    )
                    .drop(columns=["_sess_sort"], errors="ignore")
                )

            out_bytes = build_styled_completed_workbook(st.session_state.original_bytes, plan_df_export)

            st.download_button(
                "Export Completed Schedule",
                data=out_bytes,
                file_name=f"flowboard_completed_{week_start.isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.caption("Upload Excel to enable styled export.")

    active_day_list = [d for d in WEEKDAYS if plan["active_days"].get(d, False)]
    day_cols = st.columns(len(active_day_list)) if active_day_list else []

    for idx, d in enumerate(active_day_list):
        with day_cols[idx]:
            sessions = plan["day_sessions"][d]
            focus = plan["day_focus"].get(d, "(auto)")
            st.markdown(f"### {d}")
            st.caption(f"Focus: {focus}")

            if st.button("Reset", key=f"reset_{d}", use_container_width=True):
                for sess in ["AM", "PM"]:
                    plan["remaining"] = plan["buckets"][d][sess] + plan["remaining"]
                    plan["buckets"][d][sess] = []
                st.session_state.plan = plan
                st.experimental_rerun()

            for sess in ["AM", "PM"]:
                if not sessions[sess]["enabled"]:
                    continue

                st.markdown(f"**{sess}**")
                items = plan["buckets"][d][sess]

                box = st.container(border=True)
                with box:
                    if not items:
                        st.caption("— empty —")
                    else:
                        for i, job in enumerate(items, start=1):
                            job["_planned_seq"] = i
                        for job in items:
                            st.markdown(render_job(job), unsafe_allow_html=True)
                st.write("")
